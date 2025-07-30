from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def write_excel_with_chart(df, metadata, output_file):
    wb = Workbook()

    # Sheet 1: Vulnerability Results
    ws_data = wb.active
    ws_data.title = "Vulnerability Results"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Autosize columns
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_data.column_dimensions[col[0].column_letter].width = max_len + 2

    # Freeze top row
    ws_data.freeze_panes = "A2"

    # Severity-based cell formatting
    severity_fills = {
        "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "High": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "Medium": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "Low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "None": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    }

    header = [cell.value for cell in ws_data[1]]
    try:
        severity_col_idx = header.index("severity") + 1
    except ValueError:
        severity_col_idx = None

    if severity_col_idx:
        for row in ws_data.iter_rows(min_row=2, min_col=severity_col_idx, max_col=severity_col_idx):
            cell = row[0]
            if cell.value:
                severity = str(cell.value).strip().capitalize()
                fill = severity_fills.get(severity)
                if fill:
                    cell.fill = fill

    # Sheet 2: Threat Chart
    ws_chart = wb.create_sheet(title="Threat Chart")
    threat_counts = df["threat"].value_counts().reset_index()
    threat_counts.columns = ["Threat", "Count"]

    for i, col in enumerate(threat_counts.columns, start=1):
        ws_chart.cell(row=1, column=i, value=col).font = Font(bold=True)

    for row_idx, row in enumerate(threat_counts.itertuples(index=False), start=2):
        ws_chart.cell(row=row_idx, column=1, value=row.Threat)
        ws_chart.cell(row=row_idx, column=2, value=row.Count)

    chart = BarChart()
    chart.title = "Threats by Frequency"
    chart.x_axis.title = "Threat Level"
    chart.y_axis.title = "Number of Ports"
    chart.style = 12

    data_ref = Reference(ws_chart, min_col=2, min_row=1, max_row=1 + len(threat_counts))
    cats_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=1 + len(threat_counts))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_chart.add_chart(chart, "E2")

    # Sheet 3: Metadata Summary
    ws_meta = wb.create_sheet(title="Summary")
    ws_meta["A1"] = "Field"
    ws_meta["B1"] = "Value"
    ws_meta["A1"].font = ws_meta["B1"].font = Font(bold=True)

    for row_idx, (key, value) in enumerate(metadata.items(), start=2):
        ws_meta.cell(row=row_idx, column=1, value=key)
        ws_meta.cell(row=row_idx, column=2, value=value)



    severity_counts = df["severity"].value_counts()
    ws_meta.append(["", ""])  # spacer
    ws_meta.append(["Severity", "Count"])
    for sev, count in severity_counts.items():
        ws_meta.append([sev, count])


    # Sheet 5: Vulnerability Types
    if "vuln_name" in df.columns:
        vuln_counts = df["vuln_name"].value_counts().reset_index()
        vuln_counts.columns = ["Vulnerability Name", "Count"]

        ws_vuln_types = wb.create_sheet(title="Vulnerability Types")

        # Write headers
        for i, col in enumerate(vuln_counts.columns, start=1):
            ws_vuln_types.cell(row=1, column=i, value=col).font = Font(bold=True)

        # Write data rows
        for r_idx, row in enumerate(vuln_counts.itertuples(index=False), start=2):
            ws_vuln_types.cell(row=r_idx, column=1, value=row[0])
            ws_vuln_types.cell(row=r_idx, column=2, value=row[1])


        # Optional: Add bar chart
        vuln_chart = BarChart()
        vuln_chart.title = "Top Vulnerability Types"
        vuln_chart.x_axis.title = "Vulnerability Name"
        vuln_chart.y_axis.title = "Occurrences"
        vuln_chart.style = 13

        data_ref = Reference(ws_vuln_types, min_col=2, min_row=1, max_row=len(vuln_counts)+1)
        cats_ref = Reference(ws_vuln_types, min_col=1, min_row=2, max_row=len(vuln_counts)+1)
        vuln_chart.add_data(data_ref, titles_from_data=True)
        vuln_chart.set_categories(cats_ref)
        ws_vuln_types.add_chart(vuln_chart, "E2")


    wb.save(output_file)